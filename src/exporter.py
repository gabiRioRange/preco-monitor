import pandas as pd
import os
from openpyxl import load_workbook
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import logger

def salvar_excel(df, caminho_arquivo):
    """
    Salva o DataFrame em Excel. Se o arquivo já existir, adiciona as novas linhas
    usando a lógica moderna do Pandas.
    """
    if df.empty:
        logger.warning("⚠️ Sem dados para salvar.")
        return

    # Cria diretório 'data' se não existir
    pasta = os.path.dirname(caminho_arquivo)
    if pasta and not os.path.exists(pasta):
        os.makedirs(pasta)

    if not os.path.exists(caminho_arquivo):
        # Se não existe, cria novo normalmente
        df.to_excel(caminho_arquivo, index=False, engine='openpyxl')
        logger.info(f"✅ Arquivo criado: {caminho_arquivo}")
    else:
        # Se existe, precisamos descobrir onde termina a planilha atual
        try:
            # 1. Lemos o arquivo para saber quantas linhas já tem
            wb = load_workbook(caminho_arquivo)
            # Assume que os dados estão na primeira aba ativa
            sheet_name = wb.sheetnames[0] 
            start_row = wb[sheet_name].max_row
            wb.close() # Fecha para liberar o arquivo
            
            # 2. Usamos o modo 'append' ('a') com overlay
            with pd.ExcelWriter(caminho_arquivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Escrevemos os dados SEM cabeçalho (header=False)
                # Começando na linha logo abaixo da última (startrow=start_row)
                df.to_excel(writer, index=False, header=False, startrow=start_row)
                
            logger.info(f"✅ Dados adicionados ao arquivo existente: {caminho_arquivo}")
            
        except Exception as e:
            logger.error(f"❌ Erro ao tentar anexar dados: {e}")
            # Backup de segurança caso dê erro no arquivo
            backup_name = caminho_arquivo.replace(".xlsx", "_novo.xlsx")
            df.to_excel(backup_name, index=False)
            logger.warning(f"⚠️ Salvo em arquivo novo por segurança: {backup_name}")